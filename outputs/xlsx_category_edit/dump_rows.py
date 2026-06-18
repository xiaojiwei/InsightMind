from openpyxl import load_workbook

input_path = "/Users/xiao/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/wxid_98sc7godnpaq21_4ed5/msg/file/2026-06/IT人才盘点分类补充版_保留原分类.xlsx"
wb = load_workbook(input_path)
ws = wb.active

current_domain = None
for row in range(1, ws.max_row + 1):
    domain = ws.cell(row=row, column=1).value
    if domain:
        current_domain = domain
    cert = ws.cell(row=row, column=2).value
    org = ws.cell(row=row, column=3).value
    role = ws.cell(row=row, column=4).value
    if any([domain, cert, org, role]):
        print(f"{row}\t{current_domain or ''}\t{cert or ''}\t{org or ''}\t{role or ''}")
