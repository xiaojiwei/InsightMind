from openpyxl import load_workbook

output_path = "/Users/xiao/InsightMind/outputs/xlsx_category_edit/IT人才盘点分类补充版_已补充类别.xlsx"
wb = load_workbook(output_path, data_only=False)
ws = wb.active

print("sheet", ws.title)
print("max_row", ws.max_row, "max_column", ws.max_column)
print("headers", [ws.cell(1, col).value for col in range(1, 7)])
for row in range(2, 21):
    print(row, ws.cell(row, 2).value, "=>", ws.cell(row, 5).value)

assert ws.cell(1, 5).value == "类别"
assert ws.cell(1, 6).value == "持证人数"
assert [ws.cell(row, 5).value for row in range(2, 12)] == [
    "Java",
    "Java",
    "Python",
    "Python",
    "C++",
    "C++",
    "Azure",
    "AI大模型",
    "SQL",
    "CRM",
]
