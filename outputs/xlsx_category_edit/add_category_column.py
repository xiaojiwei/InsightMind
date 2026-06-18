from copy import copy
from pathlib import Path

from openpyxl import load_workbook


INPUT_PATH = Path(
    "/Users/xiao/Library/Containers/com.tencent.xinWeChat/Data/Documents/"
    "xwechat_files/wxid_98sc7godnpaq21_4ed5/msg/file/2026-06/"
    "IT人才盘点分类补充版_保留原分类.xlsx"
)
OUTPUT_PATH = Path("/Users/xiao/InsightMind/outputs/xlsx_category_edit/IT人才盘点分类补充版_已补充类别.xlsx")


def contains(text, *terms):
    lowered = text.lower()
    return any(term.lower() in lowered for term in terms)


def category_for(row_number, domain, cert, org, role):
    text = " ".join(str(part or "") for part in (cert, org, role))

    if row_number in (1, 141, 142):
        return ""

    if contains(text, "java"):
        return "Java"
    if contains(text, "python"):
        return "Python"
    if contains(text, "c++"):
        return "C++"
    if contains(text, "android"):
        return "Android"
    if contains(text, "front-end", "前端", "web应用"):
        return "前端"
    if contains(text, "salesforce", "crm"):
        return "CRM"
    if contains(text, "power platform", "低代码"):
        return "低代码"
    if contains(text, "uipath", "rpa"):
        return "RPA"
    if contains(text, "mysql 8.0 database developer", "sql开发"):
        return "SQL"

    if contains(text, "ai", "machine learning", "机器学习", "大模型", "genai", "数据科学"):
        return "AI大模型"

    if contains(
        text,
        "database administrator",
        "dba",
        "postgresql dba",
        "mongodb dba",
        "cassandra dba",
        "couchbase管理员",
        "数据库管理员",
    ):
        return "DBA"

    if contains(text, "snowflake", "云数仓"):
        return "云数仓"
    if contains(text, "mongodb"):
        return "MongoDB"
    if contains(text, "cassandra"):
        return "Cassandra"
    if contains(text, "couchbase"):
        return "Couchbase"
    if contains(text, "data engineer", "数据工程", "湖仓", "spark"):
        return "数据工程"
    if contains(text, "data management", "cdmp", "数据治理", "数据资产"):
        return "数据治理"
    if contains(text, "power bi", "tableau", "bi工程师", "数据分析师", "可视化"):
        return "BI"

    if contains(text, "devnet", "网络自动化"):
        return "网络自动化"
    if contains(text, "wireless", "无线"):
        return "无线网络"
    if contains(text, "network", "ccna", "ccnp", "ccie", "网络"):
        return "网络"

    if contains(text, "cloud security", "aws certified security", "azure security", "云安全"):
        return "云安全"
    if contains(text, "kubernetes security", "cks", "k8s安全"):
        return "Kubernetes安全"
    if contains(text, "oscp", "ethical hacker", "ceh", "渗透", "攻防", "红队"):
        return "攻防"
    if contains(text, "privacy", "数据安全", "隐私"):
        return "数据安全"
    if contains(text, "27001", "合规"):
        return "合规"
    if contains(text, "cisa", "审计"):
        return "IT审计"
    if contains(text, "cissp", "cism", "security", "安全"):
        return "信息安全"

    if contains(text, "aws"):
        return "AWS"
    if contains(text, "azure"):
        return "Azure"
    if contains(text, "google cloud", "gcp"):
        return "GCP"
    if contains(text, "finops"):
        return "FinOps"

    if contains(text, "red hat", "linux", "lpic", "rhce", "rhca", "rhcsa"):
        return "Linux"

    if contains(text, "pmi-pba", "business analysis", "cbap", "业务分析", "需求分析"):
        return "业务分析"
    if contains(text, "risk management", "风险"):
        return "风险管理"
    if contains(text, "pmp", "capm", "pgmp", "pfmp", "prince2", "项目"):
        return "项目管理"

    if contains(text, "servicenow"):
        return "ServiceNow"
    if contains(text, "itil"):
        return "ITIL"
    if contains(text, "cobit", "it治理", "内控"):
        return "IT治理"
    if contains(text, "verism", "it运营", "服务管理"):
        return "IT服务管理"

    if contains(text, "performance testing", "性能测试"):
        return "性能测试"
    if contains(text, "security tester", "安全测试"):
        return "安全测试"
    if contains(text, "mobile application testing", "移动端测试"):
        return "移动测试"
    if contains(text, "usability testing", "用户体验", "可用性"):
        return "用户体验"
    if contains(text, "istqb", "测试", "qa"):
        return "软件测试"

    if contains(text, "kubernetes", "cka", "ckad", "cncf"):
        return "Kubernetes"
    if contains(text, "devops", "sre", "terraform", "gitlab", "ci/cd", "iac"):
        return "DevOps"
    if contains(text, "scrum", "agile", "敏捷"):
        return "敏捷"

    if contains(text, "archimate"):
        return "架构建模"
    if contains(text, "togaf", "企业架构"):
        return "企业架构"

    if contains(text, "sap", "erp"):
        return "ERP"
    if contains(text, "collaboration", "microsoft 365", "协作平台", "办公平台"):
        return "协作办公"

    return str(domain or "").strip()


def copy_cell_style(source, target):
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    if source.protection:
        target.protection = copy(source.protection)
    if source.alignment:
        target.alignment = copy(source.alignment)
    if source.font:
        target.font = copy(source.font)
    if source.fill:
        target.fill = copy(source.fill)
    if source.border:
        target.border = copy(source.border)


def main():
    wb = load_workbook(INPUT_PATH)
    ws = wb.active

    ws.insert_cols(5)
    ws.column_dimensions["E"].width = max(ws.column_dimensions["F"].width or 12, 12)

    current_domain = None
    for row in range(1, ws.max_row + 1):
        old_e_cell = ws.cell(row=row, column=6)
        category_cell = ws.cell(row=row, column=5)
        copy_cell_style(old_e_cell, category_cell)

        domain = ws.cell(row=row, column=1).value
        if domain:
            current_domain = domain
        cert = ws.cell(row=row, column=2).value
        org = ws.cell(row=row, column=3).value
        role = ws.cell(row=row, column=4).value

        if row == 1:
            category_cell.value = "类别"
        else:
            category_cell.value = category_for(row, current_domain, cert, org, role)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(OUTPUT_PATH)


if __name__ == "__main__":
    main()
